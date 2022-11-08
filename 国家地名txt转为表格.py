from openpyxl import Workbook


def US2xls(filename, xlsname):
    """
    :文本转换成xls的函数
    :param filename txt文本文件名称、
    :param xlsname 表示转换后的excel文件名
    """
    print(filename)
    num = 1
    try:
        f = open(filename, encoding='utf-8')
        xls = Workbook()
        # 生成excel的方法，声明excel
        sheet = xls.get_sheet_by_name('Sheet')  # 通过表名获取
        x = 1
        while True:
            # 按行循环，读取文本文件
            line = f.readline()
            if not line:
                break  # 如果没有内容，则退出循环
            for i in range(len(line.split('\t'))):
                item = line.split('\t')[i]
                sheet.cell(x, i + 1).value = item
            x += 1  # excel另起一行
            if x == 1048576:
                xls.save(xlsname)  # 保存xls文件
                xls = Workbook()  # 创建一个新的excel文件
                sheet = xls.get_sheet_by_name('Sheet')  # 通过表名获取
                xlsname = str(xlsname).replace('.xls', str(num) + '.xls')  # 更新保存的路径
                num += 1
                x = 1  # 重置行数
        f.close()
        xls.save(xlsname)  # 保存xls文件
    except:
        raise


filename = r"F:\mxy\department\国遥\数据\geonames\allCountries\allCountries.txt"
xlsname = r"F:\mxy\department\国遥\数据\geonames\1.xlsx"
US2xls(filename, xlsname)
