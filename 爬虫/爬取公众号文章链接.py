# -*- codeing = utf-8 -*-
# @Time :2023/1/18 10:52
# @Author :xming
# @Version :1.0
# @Descriptioon :
# @Link : https://baijiahao.baidu.com/s?id=1746310400676849013&wfr=spider&for=pc
# @File :  爬取公众号文章.py

import requests
from openpyxl import Workbook

url = "https://mp.weixin.qq.com/cgi-bin/appmsg"
# 自己登录的cookies
Cookie = "pgv_pvid=913815611; fqm_pvqid=d4995edb-5e01-468d-9d77-b11c93e4e2e9; tvfe_boss_uuid=c0ca2bf658594a47; RK=P3+BSBWlQx; ptcz=2b0117eef95d8d1467b3eed632964f531396e2c1b4b28ae8b90bd6921b786f9e; ua_id=2RO1ljL2ylAdRqK5AAAAAOkU5AXPRMALJ901H53pu-M=; mm_lang=zh_CN; pac_uid=0_ffe5fa68bc796; iip=0; rewardsn=; wxtokenkey=777; wxuin=2433329500; devicetype=Windows11x64; version=63080029; lang=zh_CN; pass_ticket=BrBQN1Ega2ktwbXmdUbm7IcGUI+iicHG/KU7vBhzZVIQuG034RRoEPMqAf7974Bfq3l7IoqyaQ/1Thn5xbPPxg==; wap_sid2=CNzSpogJEooBeV9IRVZHMHRRb01kVmhEdTRyZjdDRzZYc1ZJekMyUXd2SDNkOC1oS2xrNU5za1Y1bHJBdFRtTlp1X0NUblRzUG5lVlNUbHJyeWpIcWVpaVBHVFNxb1RMZmxkNENKNEdEMmh4RjZUWERZN3RqYjBlUFpVWUd6Z0JlQ1BvbVF4UGl4dDRZNFNBQUF+MIavnZ4GOA1AAQ==; appmsg_token=1200_xoT4jWoeMzqfQ%2FzTnNCIKRlPLeAo12ZIMGmZR_scD-GImgzhjAAT0Kn4G4YluZSGiWxVWjbtbjXwQORV; uuid=ea1e1306bdf1c4e37205fb3819387bc3; rand_info=CAESIA+p/Sshn2HnbTreKaLYBs/zmVwfR6o3lHjNzLSX6l6R; slave_bizuin=3918279869; data_bizuin=3918279869; bizuin=3918279869; data_ticket=h5kSDEmJdVOvSomdLDaUEZD+OFZ+RCsiS5sV6xGmcaNZQyLfXN2A21w+4QgKrKkR; slave_sid=a0ZwV0hoY0xQMzZWeFo1STF5WTFvcGFNaW1KWVA3anpabGRmaThEQ0NaN2FvTDd0ZW02ZVMzNUt5N3pSY0ZQU2w0bWM2YUJvV1dDUzZXQmtXRlhSUk1jY1F6QUE2Q1VZUmFYTmhIYmhSVTBxUWlyOU5jQXRZU1VtazBxZHJ5cnJrTzRVUDJ6cnhpeEtWSkt2; slave_user=gh_78c9073e33c5; xid=58e37fae2db394f095a9fead99e63705; cert=ZZcGIfegCMybLOCVWDAHZRPef25TD59t"
headers = {
    "Cookie": Cookie,
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.3",
}
token = "1389188398"  # 自己公众号的token
fakeid = "3918279869%3D%3D"  # 公众号对应的id
type = '9'
data1 = {
    "token": token,
    "lang": "zh_CN",
    "f": "json",
    "ajax": "1",
    "action": "list_ex",
    "begin": "0",
    "count": "4",
    "query": "",
    "fakeid": fakeid,
    "type": type,
}


def getAllInfo(url):
    # 拿一页，存一页
    messageAllInfo = []
    # begin 从0开始
    for i in range(33):  # 设置爬虫页码
        begin = i * 4
        data1["begin"] = begin
        requests.packages.urllib3.disable_warnings()
        content_json = requests.get(url, headers=headers, params=data1, verify=False).json()
        # time.sleep(random.randint(1, 10))
        if "app_msg_list" in content_json:
            for item in content_json["app_msg_list"]:
                info = {
                    "title": item['title'],
                    "url": item['link'],
                }
                messageAllInfo.append(info)
    return messageAllInfo


def main():
    f = Workbook()  # 创建一个workbook 设置编码
    sheet = f.active  # 创建sheet表单
    # 写入表头
    sheet.cell(row=1, column=1).value = 'title'  # 第一行第一列
    sheet.cell(row=1, column=2).value = 'url'
    messageAllInfo = getAllInfo(url)  # 获取信息
    print(messageAllInfo)
    print(len(messageAllInfo))  # 输出列表长度
    # 写内容
    for i in range(1, len(messageAllInfo) + 1):
        sheet.cell(row=i + 1, column=1).value = messageAllInfo[i - 1]['title']
        sheet.cell(row=i + 1, column=2).value = messageAllInfo[i - 1]['url']
    f.save(u'公众号.xls')  # 保存文件


if __name__ == '__main__':
    main()
